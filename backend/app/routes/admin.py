from flask import Blueprint, request, jsonify, current_app
from app.models import db, Application, ArchivedApplication, Admin, BlockedDate, SlotConfig, OpenedDate, AllowedMonth
from app.config import TIME_SLOTS, DEFAULT_MAX_CAPACITY
from app.extensions import limiter
from app.services.archive import archive_expired
from werkzeug.security import check_password_hash
from datetime import datetime, timezone, timedelta, date as date_type
import jwt

admin_bp = Blueprint("admin", __name__)


def verify_token(req):
    """Проверяет JWT-токен из заголовка Authorization.

    Принимает как 'Bearer <token>', так и просто '<token>' для совместимости.
    Возвращает payload словарь при успехе, None при ошибке.
    """
    auth_header = req.headers.get("Authorization", "")
    if not auth_header:
        return None

    # Поддерживаем оба формата: "Bearer <token>" и просто "<token>"
    token = auth_header.removeprefix("Bearer ").strip()
    if not token:
        return None

    try:
        payload = jwt.decode(
            token,
            current_app.config["SECRET_KEY"],
            algorithms=["HS256"],
        )
        # Явно проверяем поле user — нет поля = невалидный токен
        if not payload.get("user"):
            return None
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


def _auth_error():
    return jsonify({"message": "Нет доступа"}), 401


@admin_bp.route("/login", methods=["POST"])
@limiter.limit("5 per minute")
def login():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"message": "Неверный логин или пароль"}), 401

    admin = Admin.query.filter_by(username=username).first()
    if admin and check_password_hash(admin.password_hash, password):
        archive_expired()
        token = jwt.encode(
            {
                "user": admin.username,
                "exp": datetime.now(timezone.utc) + timedelta(hours=24),
            },
            current_app.config["SECRET_KEY"],
            algorithm="HS256",
        )
        return jsonify({"token": token})
    return jsonify({"message": "Неверный логин или пароль"}), 401


@admin_bp.route("/applications", methods=["GET"])
def get_applications():
    if not verify_token(request):
        return _auth_error()

    search   = (request.args.get("search", "") or "").strip()
    # Защита от слишком длинных поисковых запросов
    search   = search[:200]
    sort     = request.args.get("sort", "registration_date")
    order    = request.args.get("order", "asc")
    try:
        page      = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(200, int(request.args.get("page_size", 25))))
    except (TypeError, ValueError):
        page, page_size = 1, 25

    q = Application.query

    if search:
        like = f"%{search}%"
        q = q.filter(
            db.or_(
                Application.fio.ilike(like),
                Application.phone.ilike(like),
                Application.email.ilike(like),
                Application.registration_date.ilike(like),
                Application.registration_time.ilike(like),
            )
        )

    total = q.count()

    SORT_COLUMNS = {
        "fio":               Application.fio,
        "registration_date": Application.registration_date,
        "registration_time": Application.registration_time,
        "created_at":        Application.created_at,
    }
    col = SORT_COLUMNS.get(sort, Application.registration_date)
    if order == "desc":
        q = q.order_by(col.desc())
    else:
        q = q.order_by(col.asc())

    if sort == "registration_date":
        q = q.order_by(
            Application.registration_time.asc()
            if order == "asc"
            else Application.registration_time.desc()
        )
    elif sort == "registration_time":
        q = q.order_by(
            Application.registration_date.asc()
            if order == "asc"
            else Application.registration_date.desc()
        )

    apps = q.offset((page - 1) * page_size).limit(page_size).all()

    result = [
        {
            "id": a.id,
            "fio": a.fio,
            "phone": a.phone,
            "email": a.email,
            "registration_date": a.registration_date,
            "registration_time": a.registration_time,
            "created_at": a.created_at.isoformat(),
            "status": a.status,
        }
        for a in apps
    ]

    return jsonify({
        "items":      result,
        "total":      total,
        "page":       page,
        "page_size":  page_size,
        "total_pages": max(1, -(-total // page_size)),
    })


@admin_bp.route("/applications/<int:app_id>", methods=["DELETE"])
def delete_application(app_id):
    if not verify_token(request):
        return _auth_error()
    app = Application.query.get(app_id)
    if not app:
        return jsonify({"error": "Запись не найдена"}), 404
    db.session.delete(app)
    db.session.commit()
    return jsonify({"deleted": True, "id": app_id})


@admin_bp.route("/applications/<int:app_id>/status", methods=["PATCH"])
def update_application_status(app_id):
    if not verify_token(request):
        return _auth_error()
    data = request.json or {}
    new_status = data.get("status")
    if new_status not in ("pending", "confirmed", "rejected"):
        return jsonify({"error": "Недопустимый статус"}), 400
    app = Application.query.get(app_id)
    if not app:
        return jsonify({"error": "Запись не найдена"}), 404
    app.status = new_status
    db.session.commit()
    return jsonify({"id": app_id, "status": app.status})


@admin_bp.route("/archive", methods=["GET"])
def get_archive():
    if not verify_token(request):
        return _auth_error()

    search = (request.args.get("search", "") or "").strip()[:200]
    sort = request.args.get("sort", "registration_date")
    order = request.args.get("order", "desc")
    try:
        page = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(200, int(request.args.get("page_size", 25))))
    except (TypeError, ValueError):
        page, page_size = 1, 25

    q = ArchivedApplication.query

    if search:
        like = f"%{search}%"
        q = q.filter(
            db.or_(
                ArchivedApplication.fio.ilike(like),
                ArchivedApplication.phone.ilike(like),
                ArchivedApplication.email.ilike(like),
                ArchivedApplication.registration_date.ilike(like),
            )
        )

    total = q.count()

    SORT_COLUMNS = {
        "fio": ArchivedApplication.fio,
        "registration_date": ArchivedApplication.registration_date,
        "registration_time": ArchivedApplication.registration_time,
        "archived_at": ArchivedApplication.archived_at,
    }
    col = SORT_COLUMNS.get(sort, ArchivedApplication.registration_date)
    q = q.order_by(col.desc() if order == "desc" else col.asc())

    items = q.offset((page - 1) * page_size).limit(page_size).all()

    return jsonify({
        "items": [
            {
                "id": a.id,
                "original_id": a.original_id,
                "fio": a.fio,
                "phone": a.phone,
                "email": a.email,
                "registration_date": a.registration_date,
                "registration_time": a.registration_time,
                "status": a.status,
                "created_at": a.created_at.isoformat(),
                "archived_at": a.archived_at.isoformat(),
            }
            for a in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, -(-total // page_size)),
    })


@admin_bp.route("/archive/run", methods=["POST"])
def run_archive():
    if not verify_token(request):
        return _auth_error()
    count = archive_expired()
    return jsonify({"archived": count})


@admin_bp.route("/blocked-dates", methods=["GET"])
def get_blocked_dates():
    if not verify_token(request):
        return _auth_error()
    blocked = BlockedDate.query.all()
    opened = OpenedDate.query.all()
    return jsonify({
        "blocked_dates": [b.date for b in blocked],
        "opened_weekends": [o.date for o in opened],
    })


@admin_bp.route("/toggle-weekend", methods=["POST"])
def toggle_weekend():
    if not verify_token(request):
        return _auth_error()
    data = request.json or {}
    date = data.get("date")
    if not date:
        return jsonify({"error": "Дата не указана"}), 400

    try:
        parsed = date_type.fromisoformat(date)
    except ValueError:
        return jsonify({"error": "Некорректный формат даты"}), 400

    if parsed.weekday() < 5:
        return jsonify({"error": "Это не выходной день"}), 400

    existing = OpenedDate.query.filter_by(date=date).first()
    if existing:
        db.session.delete(existing)
        db.session.commit()
        return jsonify({"opened": False, "date": date})
    else:
        db.session.add(OpenedDate(date=date))
        db.session.commit()
        return jsonify({"opened": True, "date": date})


@admin_bp.route("/toggle-date", methods=["POST"])
def toggle_date():
    if not verify_token(request):
        return _auth_error()
    data = request.json or {}
    date = data.get("date")
    if not date:
        return jsonify({"error": "Дата не указана"}), 400

    try:
        date_type.fromisoformat(date)
    except ValueError:
        return jsonify({"error": "Некорректный формат даты"}), 400

    existing = BlockedDate.query.filter_by(date=date).first()
    if existing:
        db.session.delete(existing)
        db.session.commit()
        return jsonify({"blocked": False, "date": date})
    else:
        new_block = BlockedDate(date=date)
        db.session.add(new_block)
        db.session.commit()
        return jsonify({"blocked": True, "date": date})


@admin_bp.route("/slot-configs/<date>", methods=["GET"])
def get_slot_configs(date):
    if not verify_token(request):
        return _auth_error()

    try:
        date_type.fromisoformat(date)
    except ValueError:
        return jsonify({"error": "Некорректный формат даты"}), 400

    apps = Application.query.filter_by(registration_date=date).all()
    occupied: dict[str, int] = {}
    for a in apps:
        occupied[a.registration_time] = occupied.get(a.registration_time, 0) + 1

    slot_configs = SlotConfig.query.filter_by(date=date).all()
    config_map = {sc.time: sc for sc in slot_configs}

    result = {}
    for t in TIME_SLOTS:
        cfg = config_map.get(t)
        result[t] = {
            "max_capacity": cfg.max_capacity if cfg else DEFAULT_MAX_CAPACITY,
            "is_blocked": cfg.is_blocked if cfg else False,
            "occupied": occupied.get(t, 0),
        }
    return jsonify(result)


@admin_bp.route("/update-slot", methods=["POST"])
def update_slot():
    if not verify_token(request):
        return _auth_error()
    data = request.json or {}
    date = data.get("date")
    time = data.get("time")
    max_capacity = data.get("max_capacity")
    is_blocked = data.get("is_blocked")

    if not date or not time:
        return jsonify({"error": "Не указана дата или время"}), 400

    try:
        date_type.fromisoformat(date)
    except ValueError:
        return jsonify({"error": "Некорректный формат даты"}), 400

    if time not in TIME_SLOTS:
        return jsonify({"error": "Недопустимое время"}), 400

    cfg = SlotConfig.query.filter_by(date=date, time=time).first()
    if not cfg:
        cfg = SlotConfig(date=date, time=time)
        db.session.add(cfg)

    if max_capacity is not None:
        new_capacity = max(1, int(max_capacity))

        current_count = Application.query.filter_by(
            registration_date=date,
            registration_time=time
        ).count()

        if new_capacity < current_count:
            return jsonify({
                "error": (
                    f"На это время уже записано {current_count} человек. "
                    f"Нельзя установить лимит меньше этого значения."
                )
            }), 400

        cfg.max_capacity = new_capacity
    if is_blocked is not None:
        cfg.is_blocked = bool(is_blocked)

    db.session.commit()
    return jsonify(
        {
            "date": date,
            "time": time,
            "max_capacity": cfg.max_capacity,
            "is_blocked": cfg.is_blocked,
        }
    )


@admin_bp.route("/allowed-months", methods=["GET"])
def get_allowed_months():
    if not verify_token(request):
        return _auth_error()
    months = AllowedMonth.query.order_by(AllowedMonth.year, AllowedMonth.month).all()
    return jsonify([{"year": m.year, "month": m.month} for m in months])


@admin_bp.route("/add-month", methods=["POST"])
def add_month():
    if not verify_token(request):
        return _auth_error()
    data = request.json or {}
    year = data.get("year")
    month = data.get("month")
    if not year or not month:
        return jsonify({"error": "Не указан год или месяц"}), 400
    try:
        year, month = int(year), int(month)
    except (TypeError, ValueError):
        return jsonify({"error": "Некорректные данные"}), 400
    if not (1 <= month <= 12) or year < 2000:
        return jsonify({"error": "Некорректный месяц или год"}), 400
    if AllowedMonth.query.filter_by(year=year, month=month).first():
        return jsonify({"error": "Этот месяц уже добавлен"}), 409
    db.session.add(AllowedMonth(year=year, month=month))
    db.session.commit()
    return jsonify({"year": year, "month": month}), 201


@admin_bp.route("/remove-month", methods=["POST"])
def remove_month():
    if not verify_token(request):
        return _auth_error()
    data = request.json or {}
    try:
        year, month = int(data.get("year")), int(data.get("month"))
    except (TypeError, ValueError):
        return jsonify({"error": "Некорректные данные"}), 400
    row = AllowedMonth.query.filter_by(year=year, month=month).first()
    if not row:
        return jsonify({"error": "Месяц не найден"}), 404
    month_prefix = f"{year}-{month:02d}"
    existing_application = Application.query.filter(
        Application.registration_date.like(f"{month_prefix}-%")
    ).first()

    if existing_application:
        return jsonify({
            "error": "В этом месяце есть актуальные записи"
        }), 409
    db.session.delete(row)
    db.session.commit()
    return jsonify({"removed": True})


# ─── Excel export ─────────────────────────────────────────────────────────────

@admin_bp.route("/applications/export/excel", methods=["GET"])
def export_applications_excel():
    """Export all (or filtered by date) applications as .xlsx"""
    if not verify_token(request):
        return _auth_error()

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    filter_date = request.args.get("date", "").strip()  # optional YYYY-MM-DD

    q = Application.query
    if filter_date:
        q = q.filter_by(registration_date=filter_date)
    q = q.order_by(Application.registration_date.asc(), Application.registration_time.asc())
    apps = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Записи"

    HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    headers = ["ID", "ФИО", "Телефон", "Email", "Дата записи", "Время", "Статус", "Дата создания"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    STATUS_RU = {"pending": "Ожидает", "confirmed": "Подтверждена", "rejected": "Отклонена"}

    for row_idx, a in enumerate(apps, 2):
        ws.cell(row=row_idx, column=1, value=a.id)
        ws.cell(row=row_idx, column=2, value=a.fio)
        ws.cell(row=row_idx, column=3, value=a.phone)
        ws.cell(row=row_idx, column=4, value=a.email)
        ws.cell(row=row_idx, column=5, value=a.registration_date)
        ws.cell(row=row_idx, column=6, value=a.registration_time)
        ws.cell(row=row_idx, column=7, value=STATUS_RU.get(a.status, a.status))
        ws.cell(row=row_idx, column=8, value=(
            a.created_at.strftime("%d.%m.%Y %H:%M") if a.created_at else ""
        ))

    # Auto-width
    col_widths = [6, 35, 18, 30, 14, 10, 16, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"записи_{filter_date}.xlsx" if filter_date else "все_записи.xlsx"
    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ─── System clock ─────────────────────────────────────────────────────────────

@admin_bp.route("/clock", methods=["GET"])
def get_clock():
    if not verify_token(request):
        return _auth_error()
    from app.models import SystemClock
    from app.clock import get_now
    clock = SystemClock.query.get(1)
    now = get_now()
    return jsonify({
        "manual_datetime": clock.manual_datetime if clock else None,
        "timezone_name": clock.timezone_name if clock else "Europe/Moscow",
        "effective_now": now.strftime("%Y-%m-%dT%H:%M:%S"),
    })


@admin_bp.route("/clock", methods=["POST"])
def set_clock():
    if not verify_token(request):
        return _auth_error()
    from app.models import SystemClock
    data = request.json or {}
    manual_dt = data.get("manual_datetime")  # ISO string or null to reset
    tz_name = data.get("timezone_name", "Europe/Moscow")

    clock = SystemClock.query.get(1)
    if not clock:
        clock = SystemClock(id=1)
        from app.models import db as _db
        _db.session.add(clock)

    if manual_dt:
        # Validate
        try:
            datetime.fromisoformat(manual_dt)
        except ValueError:
            return jsonify({"error": "Некорректный формат даты/времени"}), 400
        clock.manual_datetime = manual_dt
    else:
        clock.manual_datetime = None  # reset to server clock

    clock.timezone_name = tz_name
    clock.updated_at = datetime.now(timezone.utc)
    from app.models import db as _db
    _db.session.commit()

    from app.clock import get_now
    now = get_now()
    return jsonify({
        "manual_datetime": clock.manual_datetime,
        "timezone_name": clock.timezone_name,
        "effective_now": now.strftime("%Y-%m-%dT%H:%M:%S"),
    })


# ─── Daily limit ──────────────────────────────────────────────────────────────

@admin_bp.route("/daily-limit/<date>", methods=["GET"])
def get_daily_limit(date):
    if not verify_token(request):
        return _auth_error()
    from app.models import DailyLimit
    try:
        date_type.fromisoformat(date)
    except ValueError:
        return jsonify({"error": "Некорректный формат даты"}), 400

    row = DailyLimit.query.filter_by(date=date).first()
    count = Application.query.filter_by(registration_date=date).count()
    return jsonify({
        "date": date,
        "max_registrations": row.max_registrations if row else 0,
        "current_count": count,
    })


@admin_bp.route("/daily-limit", methods=["POST"])
def set_daily_limit():
    if not verify_token(request):
        return _auth_error()
    from app.models import DailyLimit
    data = request.json or {}
    date = data.get("date", "").strip()
    max_reg = data.get("max_registrations")

    if not date:
        return jsonify({"error": "Не указана дата"}), 400
    try:
        date_type.fromisoformat(date)
    except ValueError:
        return jsonify({"error": "Некорректный формат даты"}), 400
    try:
        max_reg = int(max_reg)
        if max_reg < 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Некорректное значение лимита"}), 400

    row = DailyLimit.query.filter_by(date=date).first()
    if not row:
        row = DailyLimit(date=date)
        db.session.add(row)
    row.max_registrations = max_reg
    db.session.commit()
    return jsonify({"date": date, "max_registrations": max_reg})
